#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_pricelist.py
==================
Extract the GISTDA satellite-imagery price list from ``Gistda_Price_List.pdf``
and export it to three files that share the same base name:

    Gistda_Price_List.csv    - flat/long table (UTF-8 BOM, Excel-friendly)
    Gistda_Price_List.xlsx   - one worksheet per category + an index sheet
    Gistda_Price_List.html   - styled single page, one table per category

The source PDF has 7 pages and several *different* table layouts (optical
archive/tasking tables, RADARSAT-2 SLC/Path columns, COSMO SkyMed polarization
tables, etc.).  pdfplumber's cell extraction is broken by the merged header
cells, so we parse the far cleaner *text* output line by line with a small
state machine.

Run:
    python build_pricelist.py
"""
from __future__ import annotations

import html
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = "Gistda_Price_List"
HERE = Path(__file__).resolve().parent
PDF_PATH = HERE / f"{BASE}.pdf"

# Unified column order used everywhere.
COLUMNS = [
    "category",
    "system",
    "item",
    "resolution",
    "polarization",
    "price_archive",
    "price_tasking",
    "unit",
    "note",
]

# --- regex building blocks --------------------------------------------------
PRICE = r"(?:[\d,]+|N/A)"
RES = r"\d+(?:\.\d+)?\s*(?:cm|m)\."                       # e.g. 30 cm. / 1.5 m.
RES_RANGE = r"\d[\d.\sx–\-]*(?:cm|m)\."                   # e.g. 3 x 3 – 5 x 5 m.
POL = r"(?:HH|HV|VH|VV)(?:[,/\s]+(?:HH|HV|VH|VV))*"       # e.g. HH, HV / VV, VH

# A price row for the common optical / TerraSAR / RADARSAT layout:
#   <name> <resolution> <price1> <price2>
RE_TWO_PRICE = re.compile(rf"^(?P<item>.+?)\s+(?P<res>{RES})\s+(?P<p1>{PRICE})\s+(?P<p2>{PRICE})$")
# GaoFen-3: <name> <res> <polarization> <price1> <price2>
RE_GAOFEN = re.compile(rf"^(?P<item>.+?)\s+(?P<res>{RES})\s+(?P<pol>{POL})\s+(?P<p1>{PRICE})\s+(?P<p2>{PRICE})$")
# COSMO SkyMed: <name> <res range> [<polarization>] <price>
RE_COSMO = re.compile(rf"^(?P<item>.+?)\s+(?P<res>{RES_RANGE})\s+(?P<pol>{POL})?\s*(?P<p1>[\d,]+)$")
# PLANETSCOPE special: PLANETSCOPE 3 m. Access+Download 180 240
RE_PLANET = re.compile(rf"^(?P<item>PLANETSCOPE)\s+(?P<res>{RES})\s+Access\+Download\s+(?P<p1>[\d,]+)\s+(?P<p2>[\d,]+)$")


def rec(category, system, item, res, pol, p1, p2, unit, note=""):
    return {
        "category": category,
        "system": system,
        "item": item.strip(),
        "resolution": res.strip(),
        "polarization": pol.strip() if pol else "",
        "price_archive": p1.strip() if p1 else "",
        "price_tasking": p2.strip() if p2 else "",
        "unit": unit,
        "note": note,
    }


def extract(pdf_path: Path) -> list[dict]:
    """Return the list of normalized price rows parsed from the PDF text."""
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            pages.append((p.extract_text() or "").splitlines())

    rows: list[dict] = []

    # ---- Page 1: Very-high-res optical (30-50 cm), unit baht/sq.km ----------
    cat = "รายละเอียดสูงมาก (Very High Resolution 30–50 ซม.)"
    unit = "บาท/ตร.กม. (THB/sq.km)"
    for ln in pages[0]:
        m = RE_TWO_PRICE.match(ln.strip())
        if m and re.search(r"[A-Za-zÀ-ÿ]", m["item"]):
            note = ""
            if m["item"].strip() == "SKYSAT":
                note = "ขั้นต่ำ 1,250 ตร.กม. (archive) / สั่งถ่ายโปรดติดต่อเจ้าหน้าที่; เข้าดูผ่าน API/Explorer"
            rows.append(rec(cat, "Optical", m["item"], m["res"], "", m["p1"], m["p2"], unit, note))

    # ---- Page 2: High-res optical (60cm-2m) + Video/Night ------------------
    cat = "รายละเอียดสูง (High Resolution 60 ซม.–2 ม.)"
    unit = "บาท/ตร.กม. (THB/sq.km)"
    for ln in pages[1]:
        s = ln.strip()
        m = RE_TWO_PRICE.match(s)
        if not (m and re.search(r"[A-Za-z]", m["item"])):
            continue
        item = m["item"].strip()
        if item == "Video Constellation":
            rows.append(rec(cat, "Optical", item, m["res"], "", m["p1"], m["p2"],
                            "บาท/30 วินาที (THB/30s)", "วิดีโอ ≤30 วินาที/ช่วง; ขั้นต่ำ 100 ตร.กม.; แนวถ่ายกว้าง ≥5 กม."))
        elif item == "Night Imaging":
            rows.append(rec(cat, "Optical", item, m["res"], "", m["p1"], m["p2"],
                            "บาท/ตร.กม. (THB/sq.km)", "ภาพกลางคืน; ขั้นต่ำ 100 ตร.กม."))
        else:
            rows.append(rec(cat, "Optical", item, m["res"], "", m["p1"], m["p2"], unit))

    # ---- Page 3: SPOT-6/7 (baht/sq.km) + THEOS/ไทยโชต (baht/scene) ---------
    cat = "รายละเอียดสูง (High Resolution — SPOT / ไทยโชต)"
    for ln in pages[2]:
        s = ln.strip()
        m = RE_TWO_PRICE.match(s)
        if not m:
            continue
        item = m["item"].strip()
        if item.startswith("SPOT"):
            rows.append(rec(cat, "Optical", item, m["res"], "", m["p1"], m["p2"],
                            "บาท/ตร.กม. (THB/sq.km)", "ขั้นต่ำ 100 ตร.กม. (archive) / 500 ตร.กม. (tasking)"))
        elif item == "ไทยโชต":
            rows.append(rec(cat, "Optical", "ไทยโชต (THEOS)", m["res"], "", m["p1"], m["p2"],
                            "บาท/ภาพ (THB/scene)", "ราคาที่ปรับ Orthorectification แล้ว 910 บาท/ภาพ"))

    # ---- Page 4: Medium-res LANDSAT (baht/scene) + PLANETSCOPE -------------
    cat = "รายละเอียดปานกลาง (Medium Resolution >2 ม.)"
    for ln in pages[3]:
        s = ln.strip()
        mp = RE_PLANET.match(s)
        if mp:
            rows.append(rec(cat, "Optical", "PLANETSCOPE", mp["res"], "", mp["p1"], mp["p2"],
                            "บาท/ตร.กม./ปี (THB/sq.km/yr)",
                            "Access+Download; คอลัมน์ที่ 2 = การติดตาม (Monitoring); ขั้นต่ำ 100 ตร.กม.; สัญญา 1 ปี"))
            continue
        m = RE_TWO_PRICE.match(s)
        if m and m["item"].strip().startswith("LANDSAT"):
            n = {"LANDSAT-5": "Level 1T, 7 Bands", "LANDSAT-7": "Level 1T, 8 Bands",
                 "LANDSAT-8": "Level 1T, 11 Bands", "LANDSAT-9": "Level 1T, 11 Bands"}.get(m["item"].strip(), "")
            rows.append(rec(cat, "Optical", m["item"], m["res"], "", m["p1"], m["p2"],
                            "บาท/ภาพ (THB/scene)", n))

    # ---- Page 5: RADARSAT-2 (SLC / Path Image), unit baht/scene -----------
    cat = "เรดาร์ RADARSAT-2 (C band)"
    unit = "บาท/ภาพ (THB/scene)"
    note = "price_archive = Single Look complex; price_tasking = Path Image"
    for ln in pages[4]:
        m = RE_TWO_PRICE.match(ln.strip())
        if m and "band" not in m["item"].lower():
            rows.append(rec(cat, "Radar", m["item"], m["res"], "", m["p1"], m["p2"], unit, note))

    # ---- Page 6: TerraSAR-X (archive/tasking) + COSMO SkyMed --------------
    cat = "เรดาร์ TerraSAR-X (X band)"
    unit = "บาท/ภาพ (THB/scene)"
    in_cosmo = False
    for ln in pages[5]:
        s = ln.strip()
        if s.startswith("COSMO SkyMed"):
            in_cosmo = True
            continue
        if not in_cosmo:
            m = RE_TWO_PRICE.match(s)
            if m and "band" not in m["item"].lower():
                rows.append(rec(cat, "Radar", m["item"], m["res"], "", m["p1"], m["p2"], unit))
        else:
            mc = RE_COSMO.match(s)
            if mc:
                rows.append(rec("เรดาร์ COSMO SkyMed (X band)", "Radar", mc["item"], mc["res"],
                                mc["pol"] or "", mc["p1"], "", unit,
                                "price_archive = New Acquisition; StripMap PingPong polarization: HH,VV หรือ HH,HV หรือ VV,VH"))

    # ---- Page 7: GaoFen-3 (polarization, archive/tasking) -----------------
    cat = "เรดาร์ GaoFen-3 (C band)"
    unit = "บาท/ภาพ (THB/scene)"
    for ln in pages[6]:
        m = RE_GAOFEN.match(ln.strip())
        if m:
            rows.append(rec(cat, "Radar", m["item"], m["res"], m["pol"], m["p1"], m["p2"], unit))

    return rows


# --- output writers ---------------------------------------------------------
def write_csv(rows: list[dict], path: Path) -> None:
    import csv
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)


def _fit_columns(ws, headers, data):
    for c, h in enumerate(headers, start=1):
        width = len(str(h))
        for r in data:
            width = max(width, len(str(r.get(h, ""))))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 60)


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    thin = Side(style="thin", color="C9C9C9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    # Index sheet
    idx = wb.active
    idx.title = "Index"
    idx.append(["GISTDA Satellite Imagery Price List"])
    idx["A1"].font = Font(bold=True, size=14)
    idx.append(["Source: https://www.gistda.or.th/download/Gistda_Price_List.pdf"])
    idx.append(["**ราคายังไม่รวมภาษีมูลค่าเพิ่ม (prices exclude VAT)"])
    idx.append([])
    idx.append(["Category", "Rows"])
    for c in (1, 2):
        cell = idx.cell(row=5, column=c)
        cell.fill, cell.font = head_fill, head_font

    # One sheet per category (drop all-empty columns for readability)
    categories = []
    for r in rows:
        if r["category"] not in categories:
            categories.append(r["category"])

    used_names = set()
    for cat in categories:
        crows = [r for r in rows if r["category"] == cat]
        cols = [c for c in COLUMNS if c != "category" and any(r[c] for r in crows)]
        # Sheet name: max 31 chars, unique, no illegal chars
        name = re.sub(r"[:\\/?*\[\]]", " ", cat)[:31].strip()
        base_name, i = name, 1
        while name in used_names or not name:
            i += 1
            name = f"{base_name[:28]}_{i}"
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        ws.append([cat])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill, cell.font, cell.border = head_fill, head_font, border
        for r in crows:
            ws.append([r[c] for c in cols])
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(cols)):
            for cell in row:
                cell.border, cell.alignment = border, wrap
        _fit_columns(ws, cols, crows)
        ws.freeze_panes = "A3"
        idx.append([cat, len(crows)])

    idx.append([])
    idx.append(["TOTAL", len(rows)])
    _fit_columns(idx, ["Category", "Rows"], [{"Category": c, "Rows": ""} for c in categories])
    wb.save(path)


def write_html(rows: list[dict], path: Path) -> None:
    categories = []
    for r in rows:
        if r["category"] not in categories:
            categories.append(r["category"])

    col_labels = {
        "system": "ระบบ", "item": "รายการ (Item)", "resolution": "รายละเอียดภาพ",
        "polarization": "Polarization", "price_archive": "ในคลัง / คอลัมน์ 1",
        "price_tasking": "สั่งถ่าย / คอลัมน์ 2", "unit": "หน่วย", "note": "หมายเหตุ",
    }

    parts = []
    for cat in categories:
        crows = [r for r in rows if r["category"] == cat]
        cols = [c for c in COLUMNS if c != "category" and any(r[c] for r in crows)]
        thead = "".join(f"<th>{html.escape(col_labels.get(c, c))}</th>" for c in cols)
        body = []
        for r in crows:
            tds = "".join(f"<td>{html.escape(str(r[c]))}</td>" for c in cols)
            body.append(f"<tr>{tds}</tr>")
        parts.append(
            f'<section><h2>{html.escape(cat)}</h2>'
            f'<div class="scroll"><table><thead><tr>{thead}</tr></thead>'
            f'<tbody>{"".join(body)}</tbody></table></div></section>'
        )

    doc = f"""<!doctype html>
<html lang="th">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>GISTDA Satellite Imagery Price List</title>
<style>
  :root {{ --bg:#f7f8fa; --card:#fff; --ink:#1a2230; --muted:#5b6472;
          --head:#1f4e78; --line:#e3e7ee; --stripe:#f2f5fa; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; padding:24px; background:var(--bg); color:var(--ink);
         font-family:"Segoe UI","Sarabun","Noto Sans Thai",Tahoma,sans-serif; line-height:1.5; }}
  header {{ max-width:1100px; margin:0 auto 20px; }}
  h1 {{ font-size:1.6rem; margin:0 0 6px; }}
  .sub {{ color:var(--muted); font-size:.9rem; }}
  .sub a {{ color:var(--head); }}
  main {{ max-width:1100px; margin:0 auto; }}
  section {{ background:var(--card); border:1px solid var(--line); border-radius:12px;
            padding:16px 18px; margin-bottom:20px; box-shadow:0 1px 2px rgba(0,0,0,.04); }}
  h2 {{ font-size:1.1rem; margin:0 0 12px; color:var(--head); }}
  .scroll {{ overflow-x:auto; }}
  table {{ border-collapse:collapse; width:100%; font-size:.9rem; }}
  th,td {{ padding:8px 10px; border:1px solid var(--line); text-align:left; vertical-align:top; }}
  thead th {{ background:var(--head); color:#fff; position:sticky; top:0; white-space:nowrap; }}
  tbody tr:nth-child(even) {{ background:var(--stripe); }}
  td:nth-child(n+5) {{ white-space:nowrap; }}
  footer {{ max-width:1100px; margin:16px auto 0; color:var(--muted); font-size:.8rem; }}
  @media (prefers-color-scheme: dark) {{
    :root {{ --bg:#0f141b; --card:#161c26; --ink:#e6eaf0; --muted:#9aa4b2;
            --head:#2a6db0; --line:#2a323f; --stripe:#1b222d; }}
  }}
</style>
</head>
<body>
<header>
  <h1>ราคาข้อมูลภาพถ่ายดาวเทียม GISTDA</h1>
  <div class="sub">GISTDA Satellite Imagery Price List &middot;
    ที่มา: <a href="https://www.gistda.or.th/download/Gistda_Price_List.pdf">Gistda_Price_List.pdf</a><br>
    **ราคาดังกล่าวยังไม่รวมภาษีมูลค่าเพิ่ม (prices exclude VAT) &middot;
    ติดต่อ ฝ่ายพัฒนาธุรกิจและการบริการ 0 2143 9593, usd@gistda.or.th</div>
</header>
<main>
{"".join(parts)}
</main>
<footer>รวมทั้งหมด {len(rows)} รายการ / {len(categories)} หมวดหมู่.
  สร้างจากไฟล์ PDF ต้นฉบับโดยอัตโนมัติ.</footer>
</body>
</html>"""
    path.write_text(doc, encoding="utf-8")


def main() -> None:
    rows = extract(PDF_PATH)
    write_csv(rows, HERE / f"{BASE}.csv")
    write_xlsx(rows, HERE / f"{BASE}.xlsx")
    write_html(rows, HERE / f"{BASE}.html")
    print(f"Parsed {len(rows)} rows.")
    cats = {}
    for r in rows:
        cats[r["category"]] = cats.get(r["category"], 0) + 1
    for c, n in cats.items():
        print(f"  {n:>3}  {c}")


if __name__ == "__main__":
    main()
